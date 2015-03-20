#!/usr/bin/python3.4

# Written By: Jesse Millwood
# Tested on: Python 3.4 Ubuntu 14.04
    # Need Python 3.4 for openpyxl
    # Colors get messed up in spreadsheets processed by openpyxl

# Disable Some Pylint Warning Messages

# pylint: python3
# pylint: disable= C0103
# pylint: disable= W0312
# pylint: disable= W0311
# pylint: disable= R0902
# pylint: disable= R0913
# pylint: disable= W0102
# pylint: disable= C0326
# pylint: disable= C0303
# pylint: disable= C0325
# pylint: disable= R0903

from openpyxl import load_workbook
import os.path
import csv
import re
import pickle

class bomComponent(object):
    '''
    Object that holds attributes describing
    electronic parts
    '''
    def __init__(self,
                 designator,
                 evalue,
                 package,
                 dkpn,
                 mfpn,
                 group,
                 qty,
                 stock):
        self.designator = designator
        self.evalue     = evalue
        self.fvalue     = convertEvalue(self.evalue)
        self.package    = package
        self.dkpn       = dkpn
        self.mfpn       = mfpn
        self.group      = group
        self.qty        = qty  # Quantity of same part on board
        self.sqty       = 0    # Quantity of time saved part has been used
        if stock == '':
            self.stock = stockStatus(self)
        else:
            self.stock = stock 
    def lookUpInfo(self):
        '''
        Look up info in the following order:
           * Local Saved past searches (serialized python object)
           * Digi-key web scraper
           * Octo-part
        Edit the parameters of the part:
           * Stock status
           * Digikey Part Number
           * Manufacturer 
           * Manufacturer Part Number
        '''
        pass


def saveParts(current_components):
    '''
    Save the parts into a serialized python file
    with the pickle module

    Keep track of stock status of part,
    Overall times used
    Most 
    '''
    currentparts = {'Components':current_components}
    savedParts = {}
    fname = 'savedparts.pickle'
    if os.path.isfile(fname):
        # savedparts file exists, load in parts and update it
        loadedparts = pickle.load(open(fname, 'rb'))
        updatedParts = updateParts(loadedparts, currentparts)
        pickle.dump(updatedParts, open(fname, 'wb'))
#        print(x.evalue for x in loadedparts['Components'])
    else:
        # savedparts file does not exist create a list of
        # parameters to save and pickle it
        savedParts = currentparts
        print('!====Saved Parts:')
#        for part in savedParts['Components']:
#            print(part.sqty)
        pickle.dump(savedParts, open(fname, 'wb'))

def updateParts(loadedparts, currentparts):
    '''
    Update parts and return
    '''
    updatedParts = {'Components':[]}
    # Combine same parts, update sqty attribute and add new parts
    for cpart in currentparts['Components']:
        # Check if current part is in loaded parts
        if True in [cpart.evalue == y.evalue and
                    cpart.mfpn == y.mfpn
                    for y in loadedparts['Components']]:
            print(cpart.evalue, 'Already Here')
        # if current part is in loaded parts increment sqty
            for lpart in loadedparts['Components']:
                if cpart.evalue == lpart.evalue and \
                   cpart.mfpn == lpart.mfpn:
                    lpart.sqty+=1
                    updatedParts['Components'].append(lpart)
        # else add to updated parts list
        else:
            print(cpart.evalue, 'Was not in loaded parts')
            updatedParts['Components'].append(cpart)

    return updatedParts

    
def stockStatus(component):
    '''
    Checks if the part is a stock part or not.
    Returns true or false
    '''
    if component:
        return False

def importEagleCSV(projectpath, projectname):
    '''
    Take information from Eagle CSV BOM
    and enter it into the BOM Excel File	
    * How to Create The CSV File from Eagle
        * Open Schematic
	* Run bom.ulp
	* Choose The Following Options:
	    * List Type: Parts
	    * List Attributes: Check
	    * Output format: CSV
    '''
    # Collect Components
    body = []
    infile = projectpath + projectname + '.csv'
    with open(infile , 'r') as f:
        reader = csv.reader(f)
        for rownum,row in enumerate(reader):
            if rownum == 0:
                header = row
            else:
                body.append(row[0].replace('"','').split(';'))


    # order and group components
    orderedBody = reorderParts(body)
    return orderedBody

def importPartsXSLX(infile, headerrow):
    '''
    Import Electronic part information from
    a *.xslx file 
    Return a list of dictionaries containing:
    [{'headername':[values in column]}, ... ]
    '''
    wb = load_workbook(filename=infile)
    sheet = wb.worksheets[0]

    rangestring = 'A'+str(headerrow)+':'+'AA'+str(sheet.get_highest_column())

    headers=[]
    for cell in sheet.range(rangestring)[0]:
        if cell.value != None:
            headers.append(cell.value)

    columncount = len(headers)

    for r in range(1,sheet.get_highest_row()):
        if sheet.cell(row=r, column=1).value != None:
            lastvalidrow = r
    columns = []
    for c in range(1,columncount):
        columns.append([sheet.cell(row=r, column=c).value \
                        for r in range(headerrow+1,lastvalidrow)])
    return dict(zip(headers,columns))
    


def writeBOM(parts, projectname):
    '''
    Take the ordered and grouped part info and
    write it to a standard BOM and save it
    '''
    StandardBOMFILE = '/home/jesse/Digi-Parser/SampleFiles/StandardBOM_style.xlsx'
    wb = load_workbook(filename=StandardBOMFILE)
    sheet = wb.get_sheet_by_name('BOM')
    STARTROW = 8
    r = STARTROW
    # Fill BOM
    for i, part in enumerate(parts):
       sheet.cell(row = r+i,column = 1).value = part.designator
       sheet.cell(row = r+i,column = 2).value = part.evalue + ' ' + part.package
       sheet.cell(row = r+i, column = 3).value = part.qty
       #sheet.cell(row = r+i, column = 5).value = part.mf
       sheet.cell(row = r+i, column = 14).value = 'Digi-Key'
       sheet.cell(row = r+i, column = 15).value = part.dkpn
       
    projectBOMname = projectname + 'BOM' + '.xlsx'
    wb.save(projectBOMname)
    saveParts(parts)
    
def convertEvalue(evalue):
    '''
    Given a string representing an 
    electrical value (10mF, 100k Ohm)
    a float value is returned
    '''
    # Take out white spaces
    evalue = str(evalue).replace(' ','')
    if len(evalue) == 0:
        return 0.0
    elif evalue[0].isdigit():
        multipliers = {'M':10**6,
                       'K':10**3,
                       'k':10**3,
                       'm':10**-3,
                       'u':10**-6,
                       'n':10**-9,
                       'p':10**-12}
        baseval = re.findall(r'\d*\.?\d+',evalue)[0]
        if len(baseval)==len(evalue): # no multiplier
            return float(baseval)
        elif evalue[len(baseval)] in multipliers.keys(): # multiplier exists
            return float(baseval)*multipliers[evalue[len(baseval)]]
        else:
            return float(baseval)
    else:
        return 0.0

def setComponentGroup(component):
    '''
    Given a row of the csv file 
    the grouping of the component will
    be extracted and returned
    '''
    # TODO: use regex to parse component info and figure out what the group is
    # TODO make easier to add a group and not have to manually shift the precedence numbers for all of the components
    # TODO crystal and zeners still not showing up correctly
    passive_groups = [('C', 'Capacitor', 0, 'Passive '),
                      ('R', 'Resistor', 1, 'Passive '),
                      ('R', 'Potentiometer', 2, 'Passive '),
                      ('D', 'Diode', 3, 'Passive '),
                      ('L', 'Inductor', 4, 'Passive '),
                      ('LED', 'LED', 5, 'Passive ')]
    transistors = [('U', 'FET', 6, 'Transistor '),
                   ('U', 'Transistor', 6, 'General '),
                   ('U', 'NPN', 7, 'Transistor '),
                   ('U', 'PNP', 8, 'Transistor ')]
    mcu = [('U', 'atmega', 9, 'IC ')]
    ic = [('U', 'Reg', 10, 'IC '),
          ('U', 'Transceiver', 11, 'IC '),
          ('U', 'Comparator', 12, 'IC '),
          ('U', 'ESD', 13, 'IC ')]
    sw = [('SW', 'Switch', 101, '')]
    crystal = [('Q', 'crystal ', 100, '')]

    # Sample Component Value:
    # ['C1', '4.7uF', 'CAPACITOR0805', 'C0805']
    # Sample Group Value:
    # ('Passive Resistor', 1)
    for label, ptype, precedence, gtype in passive_groups + \
        transistors + mcu + crystal + ic + sw:
        for parameter in filter(lambda x: ptype.lower() in str(x).lower(), component):
            return (gtype+ptype, precedence)
    # Component did not fit any of the defined groups
    # Set precedence to 1000 to keep at the end
    print(component)
    return ('Other', 1000)
    
def reorderParts(parts):
    '''
    Function: reorderParts
    Description: Groups and orders the
    components of the bill of materials
    according to value
	Order of Priority:
	* Group (Resistor, Capacitor, IC ..)
	* Stock Part (YES, NO)
	* Value (If applicable)
        * Group the same values together and record qty
    '''
    # Make list of objects containing information about
    # the components in the bom
    bomComponents = []
    for component in parts:
        #print (component)
        # Create a list of objects with the proper attributes
        if True in ['-ND' in x for x in component]:
            diginum = component[5]
        else:
            diginum = ''
        bomComponents.append(bomComponent(designator = component[0],
                                          evalue = component[1],
                                          package = component[3],
                                          group = setComponentGroup(component),
                                          dkpn = diginum,
                                          mfpn = '',
                                          qty = 1))

    # Look up and associate part info (dk part no, mf part no)
    for component in bomComponents: component.lookUpInfo()
    # Order By Group Precedence and value
    bomComponents.sort(key = lambda x: (x.group[1], x.fvalue))
    bomComponents = combineSameComponents(bomComponents)    
    return bomComponents

def combineSameComponents(body):
        '''
        Combines components with the same 
        fvalue and group attributes
        increments respective qty attribute
        for component object
        '''
        combined = []
        copy = body
        counted = []
        for i,member1 in enumerate(body):
            if i not in counted:
                designators = member1.designator
                for i, member2 in enumerate(copy):
                    if member1.fvalue == member2.fvalue and \
                       member1.evalue == member2.evalue and \
                       member1.group[0] == member2.group[0] and \
                       member1.designator != member2.designator and \
                       'Other' not in member1.group[0]:
                        counted.append(i)
                        member1.qty += 1
                        designators += ' '+ member2.designator
                member1.designator = designators
                combined.append(member1)
        for thing in combined:
            print(thing.qty, thing.designator, thing.evalue)
        return combined

def checkpackage(parameters):
    '''
    Given a list of part parameters, the parameters
    will be checked to see if they contain any strings
    matching a standard footprint
    '''

    packages = ['0402',
                '0805',
                '1007',
                '1206',
                'SIP',
                'DO204',
                'DO213','MELF',
                'DO214','SMA','SMB','SMC',
                'SOD',
#                'DIP',
                'DFN',
                'MSOP',
                'SOIC',
                'SC70','SC71',
                'SOT23','TSOT',
                'TO3','TO5','TO18','TO66','TO92',
                'TO126','TO220','TO263',
                'D2PAK',
                'DPAK',
                'SSOP',
                'TDFN',
                'TSOP',
                'SSOP',
                'TSSOP',
                'PLCC',
                'CLCC',
                'LQFP',
                'TQFP',
                'TQFN',
                'BGA',
                'PTH']
    parameters = [str(x).replace('-','') for x in parameters]

    detectedpackages = []
    for parameter in parameters:
        for package in packages:
            if package in parameter:
                detectedpackages.append(package)
            
    # Make sure only one package is detected and return it
    if len(detectedpackages) > 0:
        if all(x == detectedpackages[0] for x in detectedpackages):
            return detectedpackages[0]
        else:
            print("Multiple packages detected")
            print(detectedpackages)
            print(parameters)
            return -1
    else:
        print("No Packages Detected")
        print(parameters)
        return -1
if __name__ == '__main__':
    # Some test cases
    pinfo = ['R-0805-20k',
             'Resistor, 20k Ohm, 1%, 0805, 1/8W',
             'Yageo','RC0805FR-0720KL','0.01']
    sinfo = ['RC0805FR-073K9L',
             'R',
             'Resistor, 3.9k Ohm, 1%, 0805, 1/8W',
             'R-0805-3.9k', 'Yageo']
    ssinfo = ['RC0805JR-070RL',
              'R',
              'Resistor, Jumper, 0805, 1/8W',
              'R-0805-0', 'Yageo']



    #footprint = checkpackage(pinfo)
    sfootprint= checkpackage(sinfo)
    # if footprint != -1:
    #     print("Package: {}".format(footprint))
    # else:
    #     print("ERROR")
    ifile = '/home/jesse/BOMtools/SampleFiles/Magnum_Stock_Parts.xlsx'
    partcols = importPartsXSLX(ifile,4)
