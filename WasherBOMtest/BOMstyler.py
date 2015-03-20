#! /usr/bin/python3.4

from  openpyxl import load_workbook
from openpyxl.styles import Style, Font, colors, PatternFill,Alignment, Border, Side
# headerstyle = Style(font=Font( name='Calibre',
#                                size=11,
#                                color= colors.WHITE),
#                     fill=PatternFill( fill_type = 'solid',
#                                       fgColor= colors.BLACK),
#                     alignment=Alignment( horizontal='center',
#                                          vertical='center'))
rowstyle = Style(font=Font( name='Arial',
                            size=11,
                            color=colors.GREEN),
                 border=Border(left=Side(border_style='thin',
                                         color= colors.RED)))
if __name__ == '__main__':
    wb = load_workbook(filename='./StandardBOM_style.xlsx')
    sheet = wb.get_sheet_by_name('BOM')

    STARTCOL = 1
    ENDCOL   = 32
    HEADROW = (6,7)
    BODYROWS= {'MIN':8, 'MAX':60}

    #sheet.row(8).style=rowstyle
    #sheet.row(8).height=10
    # sheet.row_dimensions[8].height = 30
    # sheet.row_dimensions[8].style = rowstyle
    # for i in range(STARTCOL,ENDCOL+1):
    #     for headrow in HEADROW:
    #         sheet.cell(row = headrow, column = i).style = headerstyle
    
    for i in range(BODYROWS['MIN'],BODYROWS['MAX']):
        sheet.cell(row = i, column = 1).value = 'Hello'
        sheet.cell(row = i, column = 1).style = rowstyle
                
    

    wb.save('TestBOM.xlsx')





